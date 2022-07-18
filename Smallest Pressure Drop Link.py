import osmnx as ox
import networkx as nx
import wntr
import pandas as pd
import os
import openpyxl
import tifffile as tiff
import numpy as np
from osgeo import gdal
import numpy as np
import re
from openpyxl import load_workbook

# os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Simulation")
# clusterfile = 'Centralized_elevcluster' + str(1) + '.csv'

# road_df = pd.read_csv('Centralized_elevcluster1.csv')
# road_df.head()

# road_df_num=road_df.loc[:,['V1','V2']]
# wb=openpyxl.load_workbook("Centralized_elevcluster1.xlsx")
# ws=wb['Centralized_elevcluster1']

# rows=ws.iter_rows(min_row=2,max_row=1182,min_col=2,max_col=3)
# # print(rows)
# x=[]
# y=[]
# for a,b in rows:
#     x.append(a.value)
#     y.append(b.value)

# print(x)

# gets the drivable street network within some lat-long bounding box
xmin = -87.52739617137372
xmax = -87.4841138757
ymin = 32.42075991564647
ymax = 32.4687035515

G=ox.graph_from_bbox(north=ymax, south=ymin,east=xmax,west=xmin,network_type='drive_service')
# G_projected = ox.project_graph(G)
G_projected = ox.get_undirected(G)
ox.plot_graph(G_projected)
G_undirected=G_projected

Node_dict=G_undirected.nodes(data=True)
Edge_dict=G_undirected.edges(data=True)
# print(G.nodes(data=True))
# print(G.edges(data=True))
# print(Node_dict)
# print(len(Node_dict))
# print(len(Edge_dict))

Node_L=list(G_undirected.nodes)
Edge_L=list(G_undirected.edges)

# print(Node_L)
# print(Edge_L)
# Edge_L[0][1]
# print(str(Edge_L[0][1]))
# print(G.edges.values())
    
# get node coordinates with index
# node_id = list(G.nodes)[3]
# G.nodes[node_id]['x'] #lon
# G.nodes[node_id]['y'] #lat
# print(G.nodes[node_id]['x'] )
# print(G.nodes[node_id]['y'] )


# edge_id = list(G.edges)[0]
# G.edges[edge_id]['length']
# print(G.edges[edge_id]['length'])



os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Simulation")

#add an elevation atribute to the road nodes that way when we deal with them later we are chilling 
im = tiff.imread('./USGS_13_n33w088_20190918.tif')
imarray = np.array(im)
im.shape
tiff.imshow(im)

xmin = -87.52739617137372
xmax = -87.4841138757
ymin = 32.42075991564647
ymax = 32.4687035515

xdiff = xmax - xmin
ydiff = ymax - ymin
xstep = xdiff / imarray.shape[1] #remember DEM does row then column so we need to get the column length here
ystep = ydiff / imarray.shape[0] #remember DEM does row then column so we need to get the row length here

#I ensured these elevations were correct by checking three random ones off GEE
for l in G_undirected.nodes:
    if type(l) == int:
        frompointlon = G_undirected.nodes[l]['x']
        frompointlat = G_undirected.nodes[l]['y']
        p1_column = int((frompointlon - xmin) // xstep)
        #gets column for the numpy array with all elevations
        p1_row = int((ymax - frompointlat) // ystep) #remember y starts from the upper right hand corner at 0 then goes down so this is better
        G_undirected.nodes[l]['elevation'] = fromelevation = imarray[p1_row, p1_column]
#         print(G.nodes[l]['elevation'] )

# #read the elevation from osgeo import gdal
# gdal.UseExceptions()

# ds = gdal.Open('USGS_13_n33w088_20190918.tif')
# band = ds.GetRasterBand(1)
# elevation = band.ReadAsArray()

# print(elevation.shape)
# print(elevation)

# for l in G.nodes:
#     G.nodes[l]['elevation'] = elevation[l]
#     print(G.nodes[l]['elevation'] )


wn = wntr.network.WaterNetworkModel()

wn.add_pattern('pat1', [0.5,1.3,1,1.2])

for i in range(len(Node_dict)):
    node_id = list(G_undirected.nodes)[i]
    # print(node_id)
    G_undirected.nodes[node_id]['x'] #lon
    G_undirected.nodes[node_id]['y'] #lat
    # G.nodes[node_id]['elevation']
    # print(G.nodes[node_id]['x'] )
    # add_junction(name, base_demand=0.0, demand_pattern=None, elevation=0.0, coordinates=None, demand_category=None)
    wn.add_junction(str(Node_L[i]), base_demand=0.001, demand_pattern='pat1', 
                    elevation=G_undirected.nodes[node_id]['elevation'], 
                    coordinates=(G_undirected.nodes[node_id]['x'],G_undirected.nodes[node_id]['y'] ))
    # print(G.nodes[node_id]['elevation'])
    
for j in range(len(Edge_dict)):
    edge_id = list(G_undirected.edges)[j]
    # print(edge_id)
    # G.edges[edge_id]['source']
    # G.edges[edge_id]['target']
    G_undirected.edges[edge_id]['length']
    wn.add_pipe(str(j), str(Edge_L[j][0]), str(Edge_L[j][1]), 
                length=G_undirected.edges[edge_id]['length'], diameter=0.5, roughness=100, minor_loss=0.0)
    # print(G.edges[edge_id]['length'])

# find nodes with elevation greater than 100m
junction_elevation_100 = wn.query_node_attribute('elevation', np.greater_equal,
     100, node_type=wntr.network.model.Junction)
print(junction_elevation_100)


# convert the node with highest elevation to be reservoir 
wn.remove_link('341')
wn.remove_link('347')
wn.remove_link('241')
wn.remove_link('290')
wn.remove_link('339')
wn.remove_link('12')
wn.remove_node('59153773')

junction = wn.get_node('59095809')
junction.demand_timeseries_list[0].base_value = 0
junction = wn.get_node('59079200')
junction.demand_timeseries_list[0].base_value = 0
# junction = wn.get_node('59103180')
# junction.demand_timeseries_list[0].base_value = 0
# junction = wn.get_node('59096486')
# junction.demand_timeseries_list[0].base_value = 0
# junction = wn.get_node('59079187')
# junction.demand_timeseries_list[0].base_value = 0
# junction = wn.get_node('59095779')
# junction.demand_timeseries_list[0].base_value = 0
# junction = wn.get_node('59101483')
# junction.demand_timeseries_list[0].base_value = 0
# pipe = wn.get_link('279')
# pipe.diameter = pipe.diameter*1.5
# pipe = wn.get_link('327')
# pipe.diameter = pipe.diameter*2.0
# pipe = wn.get_link('310')
# pipe.diameter = pipe.diameter*2.0
# pipe = wn.get_link('322')
# pipe.diameter = pipe.diameter*2.0

wn.add_reservoir('reservoir', base_head=114.757561, head_pattern=None, coordinates=(-87.507,32.467))


edge_id = list(G.edges)[341]
G_undirected.edges[edge_id]['length']    
wn.add_pipe('341', str(Edge_L[341][0]), 'reservoir', length=G_undirected.edges[edge_id]['length'], diameter=0.3556, roughness=100,
      minor_loss=0.0)
edge_id= list(G.edges)[347]
G_undirected.edges[edge_id]['length']    
wn.add_pipe('347', str(Edge_L[347][0]), 'reservoir', length=G_undirected.edges[edge_id]['length'], diameter=0.3556, roughness=100,
      minor_loss=0.0)

ax = wntr.graphics.plot_network(wn)

# add source for chemical quality water analysis
wn.add_source('Source', 'reservoir', 'CONCEN', 0.001, pattern=None)
wn.options.quality.parameter = 'CHEMICAL'


reservoir=wn.get_node('reservoir')
reservoir.initial_quality= 0.001

# modify simulation duration
wn.options.time.duration = 240*3600
wn.options.time.pattern_timestep = 6*3600


# Simulate hydraulics
sim = wntr.sim.EpanetSimulator(wn)
results = sim.run_sim()


pressure = results.node['pressure']
pressure_avg = pressure.mean()
# print(pressure_avg['59146257'])
flowrate = results.link['flowrate']
flowrate_avg = flowrate.mean()

os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Optimization")
wb = load_workbook(filename='flowrate length.xlsx')
ws = wb.active

flowrate = {}
for column in list(ws.columns)[1:]:
    flowrate [column[1].value,column[2].value,column[3].value,column[0].value] = [(c.value) for c in column[4:]][0]

x_pos = []
x_neg = []
y_pos = []
y_neg = []
link_pos = []
link_neg = []
for k,v in flowrate.items():
    if v>0:
        x_pos.append(k[0])
        y_pos.append(k[1])
        link_pos.append(k[3])
    if v<0:
        x_neg.append(k[1])
        y_neg.append(k[0])
        link_neg.append(k[3])

x_new = x_pos + x_neg
y_new = y_pos + y_neg
link_new = link_pos + link_neg
link = dict(zip(zip(x_new, y_new),link_new))

# for key,value in link.items():
    # start = str(key[0])
    # end = str(key[1])
    # value = str(value)
    # print(pressure_avg[str(key[1])])
    # print(flowrate_avg[value])

# calculate pressure drop for each link
pressure_drop = {}    
for key,value in link.items():
    x = abs(pressure_avg[str(key[0])] - pressure_avg[str(key[1])]) / abs(flowrate_avg[value])
    pressure_drop [key] = x

# find 10% links with smallest pressure drop
sort = sorted(((v,k) for k,v in pressure_drop.items()))
link_min = []
for i in range (int(0.10 * len(sort))):
    link_min.append(sort[i][1])

# rest of links
link_res = []
for i in range (int(0.10 * len(sort)),len(sort)):
    link_res.append(sort[i][1])    

# pressure_at_5hr = results.node['pressure'].loc[5*3600, :]
# wntr.graphics.plot_network(wn, node_attribute=pressure_at_5hr, node_size=30, 
#                         title='Pressure at 5 hours')

# pressure = results.node['pressure']
# pressure_at_5hr = pressure.loc[5*3600,:]
# print('Pressure at 5 hours:')
# print(pressure_at_5hr)



# quality = results.node['quality']
# quality_at_10hr = quality.loc[10*3600,:]
# print('Quality at 10 hours:')
# print(quality_at_10hr)


# flowrate = results.link['flowrate']

# print(dict(wn.options.reaction)) 

# # create elevation dictionary
# elevation = []
# node = []
# for i in range(len(Node_dict)):
#     node_id = list(G_undirected.nodes)[i]
#     elevation.append(G_undirected.nodes[node_id]['elevation'])
#     node.append(Node_L[i])
    
# elevation = dict(zip(node,elevation))

# # create pipe length dictionary
# length = []
# link = []
# for i in range(len(Edge_dict)):
#     edge_id = list(G_undirected.edges)[i]
#     length.append(G_undirected.edges[edge_id]['length'])
#     link.append(Edge_L[i])
    
# length = dict(zip(link,length))

# os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Optimization")
# df = pd.DataFrame(length,index=[0]).T  # transpose to look just like the sheet above
# df.to_excel('length.xls')    

# os.chdir(r'C:\Users\12757\Desktop\Columbia\M.S. Thesis\WNTR-main\examples\networks')
# # saved as inp file
# wn.write_inpfile('Undirected Roadnetwork.inp', version=2.2)

# # export simulation results
# os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Optimization")
# pressure.to_excel('pressure3.xlsx')
# quality.to_excel('quality2.xlsx')
# flowrate.to_excel('flowrate2.xlsx')

# export list to excel
# pd.DataFrame(Edge_L).to_excel('Edge.xlsx', header=False, index=True)