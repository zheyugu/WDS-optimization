import osmnx as ox
import networkx as nx
import wntr
import pandas as pd
import os
import openpyxl
import tifffile as tiff
import numpy as np
from osgeo import gdal
import numpy as np
import re
from openpyxl import load_workbook
import gurobipy as gp
from gurobipy import GRB
import sys
import csv
import math

# gets the drivable street network within some lat-long bounding box
xmin = -87.52739617137372
xmax = -87.4841138757
ymin = 32.42075991564647
ymax = 32.4687035515

G=ox.graph_from_bbox(north=ymax, south=ymin,east=xmax,west=xmin,network_type='drive_service')
# G_projected = ox.project_graph(G)
G_projected = ox.get_undirected(G)
ox.plot_graph(G_projected)
G_undirected=G_projected

Node_dict=G_undirected.nodes(data=True)
Edge_dict=G_undirected.edges(data=True)

Node_L=list(G_undirected.nodes)
Edge_L=list(G_undirected.edges)

os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Simulation")
#add an elevation atribute to the road nodes that way when we deal with them later we are chilling 
im = tiff.imread('./USGS_13_n33w088_20190918.tif')
imarray = np.array(im)
im.shape
tiff.imshow(im)

xmin = -87.52739617137372
xmax = -87.4841138757
ymin = 32.42075991564647
ymax = 32.4687035515

xdiff = xmax - xmin
ydiff = ymax - ymin
xstep = xdiff / imarray.shape[1] #remember DEM does row then column so we need to get the column length here
ystep = ydiff / imarray.shape[0] #remember DEM does row then column so we need to get the row length here

#I ensured these elevations were correct by checking three random ones off GEE
for l in G_undirected.nodes:
    if type(l) == int:
        frompointlon = G_undirected.nodes[l]['x']
        frompointlat = G_undirected.nodes[l]['y']
        p1_column = int((frompointlon - xmin) // xstep)
        #gets column for the numpy array with all elevations
        p1_row = int((ymax - frompointlat) // ystep) #remember y starts from the upper right hand corner at 0 then goes down so this is better
        G_undirected.nodes[l]['elevation'] = fromelevation = imarray[p1_row, p1_column]

wn = wntr.network.WaterNetworkModel()

wn.add_pattern('pat1', [0.5,1.3,1,1.2])

for i in range(len(Node_dict)):
    node_id = list(G_undirected.nodes)[i]
    # print(node_id)
    G_undirected.nodes[node_id]['x'] #lon
    G_undirected.nodes[node_id]['y'] #lat
    # G.nodes[node_id]['elevation']
    # print(G.nodes[node_id]['x'] )
    # add_junction(name, base_demand=0.0, demand_pattern=None, elevation=0.0, coordinates=None, demand_category=None)
    wn.add_junction(str(Node_L[i]), base_demand=0.001, demand_pattern='pat1', 
                    elevation=G_undirected.nodes[node_id]['elevation'], 
                    coordinates=(G_undirected.nodes[node_id]['x'],G_undirected.nodes[node_id]['y'] ))
    # print(G.nodes[node_id]['elevation'])
    
for j in range(len(Edge_dict)):
    edge_id = list(G_undirected.edges)[j]
    # print(edge_id)
    # G.edges[edge_id]['source']
    # G.edges[edge_id]['target']
    G_undirected.edges[edge_id]['length']
    wn.add_pipe(str(j), str(Edge_L[j][0]), str(Edge_L[j][1]), 
                length=G_undirected.edges[edge_id]['length'], diameter=0.5, roughness=100, minor_loss=0.0)
    # print(G.edges[edge_id]['length'])

# find nodes with elevation greater than 100m
junction_elevation_100 = wn.query_node_attribute('elevation', np.greater_equal,
     100, node_type=wntr.network.model.Junction)
print(junction_elevation_100)

# convert the node with highest elevation to be reservoir 
wn.remove_link('341')
wn.remove_link('347')
wn.remove_link('241')
wn.remove_link('290')
wn.remove_link('339')
wn.remove_link('12')
wn.remove_node('59153773')

junction = wn.get_node('59095809')
junction.demand_timeseries_list[0].base_value = 0
junction = wn.get_node('59079200')
junction.demand_timeseries_list[0].base_value = 0

wn.add_reservoir('reservoir', base_head=114.757561, head_pattern=None, coordinates=(-87.507,32.467))

edge_id = list(G.edges)[341]
G_undirected.edges[edge_id]['length']    
wn.add_pipe('341', str(Edge_L[341][0]), 'reservoir', length=G_undirected.edges[edge_id]['length'], diameter=0.3556, roughness=100,
      minor_loss=0.0)
edge_id= list(G.edges)[347]
G_undirected.edges[edge_id]['length']    
wn.add_pipe('347', str(Edge_L[347][0]), 'reservoir', length=G_undirected.edges[edge_id]['length'], diameter=0.3556, roughness=100,
      minor_loss=0.0)

ax = wntr.graphics.plot_network(wn)

# add source for chemical quality water analysis
wn.add_source('Source', 'reservoir', 'CONCEN', 0.001, pattern=None)
wn.options.quality.parameter = 'CHEMICAL'


reservoir=wn.get_node('reservoir')
reservoir.initial_quality= 0.001

# modify simulation duration
wn.options.time.duration = 240*3600
wn.options.time.pattern_timestep = 6*3600


# Simulate hydraulics
sim = wntr.sim.EpanetSimulator(wn)
results = sim.run_sim()

# calculate average pressure and flowrate
pressure = results.node['pressure']
pressure_avg = pressure.mean()
# print(pressure_avg['59146257'])
flowrate = results.link['flowrate']
flowrate_avg = flowrate.mean()

os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Optimization")
wb = load_workbook(filename='flowrate length.xlsx')
ws = wb.active

flowrate = {}
for column in list(ws.columns)[1:]:
    flowrate [column[1].value,column[2].value,column[3].value,column[0].value] = [(c.value) for c in column[4:]][0]

x_pos = []
x_neg = []
y_pos = []
y_neg = []
v_pos = []
v_neg = []
length_pos = []
length_neg = []
link_pos = []
link_neg = []
for k,v in flowrate.items():
    if v>0:
        x_pos.append(k[0])
        y_pos.append(k[1])
        v_pos.append(v)
        length_pos.append(k[2])
        link_pos.append(k[3])
    if v<0:
        x_neg.append(k[1])
        y_neg.append(k[0])
        # v1.append(v[0])
        # new = dict((merge(x,y), -v)
        v_neg.append(-v)
        length_neg.append(k[2])
        link_neg.append(k[3])

x_new = x_pos + x_neg
y_new = y_pos + y_neg
v_new = v_pos + v_neg
link_new = link_pos + link_neg
length_new = length_pos + length_neg

new_flowrate = dict(zip(zip(x_new, y_new),v_new))
length = dict(zip(zip(x_new, y_new),length_new))
link = dict(zip(zip(x_new, y_new),link_new))

wb = load_workbook(filename='elevation.xlsx')
ws = wb.active

elevation = {}
for column in list(ws.columns)[1:]:
    elevation [column[0].value] = [(c.value) for c in column[1:]][0]

# calculate pressure drop for each link
pressure_drop = {}    
for key,value in link.items():
    x = abs(pressure_avg[str(key[0])] - pressure_avg[str(key[1])]) / abs(flowrate_avg[value])
    pressure_drop [key] = x

# find 10% links with lowest pressure drop
sort = sorted(((v,k) for k,v in pressure_drop.items()))
link_min = []
for i in range (int(0.10 * len(sort))):
    link_min.append(sort[i][1])

# rest of links
link_res = []
for i in range (int(0.10 * len(sort)),len(sort)):
    link_res.append(sort[i][1])    

# define pipe size options for each link
pipe_size = {}
for i in link_min:
    pipe_size [i] = [0.05, 0.06, 0.08, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45]
    for j in link_res:
        pipe_size[j] = [0.5]

indice = [(x[0],x[1] ,z) for x, y in pipe_size.items() for z in y]

m = gp.Model('Pipe Sizing')

link,flowrate = gp.multidict(new_flowrate)
node,elevation = gp.multidict(elevation)

# set parameter value
D1 = [0.05, 0.06, 0.08, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45]
D2 = [0.5]
Diameter = {0.05:0.05, 0.06:0.05, 0.08:0.08, 0.10:0.10, 0.15:0.15, 0.20:0.20, 
            0.25:0.25, 0.30:0.30, 0.35:0.35, 0.40:0.40, 0.45:0.45, 0.5:0.5}
L = length #Length of pipe connecting node i to node j [m]
S_min = 0.001 #minimum pipe slope allowed. 0.1% for pressurized flow
S_max = 0.1 #maximum pipe slope allowed. 10% or 0.1 m/m
P_max = 353 #maximum number of pump stations that can be placed in the model
Q = new_flowrate #pipe flow between nodes i and j [m3/s]
V_min = 0.6 #0.6 [m/s] 
V_max = 3 #3 [m/s] 
EL = elevation  #Ground elevation at node i [m] determined from the LiDAR data
# W = 0.5 #0.5 meters added trench width on each side of the pipe [m] 
# CE = 1000 #cost of excavation $/m3
# CB = 100 #cost of bedding gravel $/m2
CPS = 10 #capital of a pump station $/unit
CP = {0.05:50, 0.06:60, 0.08:80, 0.10:100, 0.15:150, 0.20:200, #material costs of piping 
      0.25:250, 0.30:300, 0.35:350, 0.40:400, 0.45:450, 0.5:500}  #  $/m of pipe of diameter k
PS_OM = 10 #Operations and maintenance cost for pump station [$/PS]
COL_OM = 1000000 #Operations and maintenance cost for collection system piping [$/connection]
N = [] #number of nodes i that contribute wastewater to the system

for i in range(290):
    N.append(2)
    
N = dict(zip(node,N))

# create variables
d = m.addVars(indice, vtype=GRB.BINARY, name="d")
e_in = m.addVars(node, vtype=GRB.CONTINUOUS, name="e_in")
e_out = m.addVars(node, vtype=GRB.CONTINUOUS, name="e_out")
p = m.addVars(link, vtype=GRB.BINARY, name="p")
pc = m.addVars(link,vtype=GRB.CONTINUOUS, name="pc")
h = m.addVars(node,vtype=GRB.CONTINUOUS, name="h")

# create constraints
m.addConstrs((gp.quicksum(d[i,j,k] for k in D1) == 1 for i,j in link_min), name='Pipe Size')
m.addConstrs((gp.quicksum(d[i,j,k] for k in D2) == 1 for i,j in link_res), name='Pipe Size')
m.addConstrs((h[i] == e_out[i]-e_in[i] for i in node), name='Elevation Change')
m.addConstrs((h[i] + sys.float_info.epsilon <= p[i,j] for i,j in link), name='Pump')
m.addConstrs((S_max >= (e_in[i]-e_in[j]+h[i])/L[i,j] for i,j in link), name='Max Slope')
m.addConstrs(((e_in[i]-e_in[j]+h[i])/L[i,j] >= S_min for i,j in link), name='Min Slope')
m.addConstrs((gp.quicksum(math.pi/4 * Diameter[k] ** 2 * d[i,j,k] for k in D1) >= 
                  Q[i,j] / V_max for i,j in link_min), name='Velocity')
m.addConstrs((gp.quicksum(math.pi/4 * Diameter[k] ** 2 * d[i,j,k] for k in D2) >= 
                  Q[i,j] / V_max for i,j in link_res), name='Velocity')
m.addConstrs((Q[i,j] * p[i,j] <= pc[i,j] for i,j in link), name='Pump Capacity')
m.addConstr((gp.quicksum(p[i,j] for i,j in link) <= P_max ), name='Max Pump Number')

# create objective
obj2 = gp.quicksum(L[i,j] * gp.quicksum(CP[k] * Diameter[k] * d[i,j,k] for k in D1) 
                    for i,j in link_min) + gp.quicksum(L[i,j] * gp.quicksum(CP[k] * 
                    Diameter[k] * d[i,j,k] for k in D2) for i,j in link_res)
obj3 = gp.quicksum(CPS * p[i,j] for i,j in link)
obj4 = gp.quicksum(COL_OM * N[i] for i in node) + gp.quicksum(PS_OM * p[i,j] 
                                                              for i,j in link)
obj = obj2 + obj3 + obj4

m.setObjective((obj),GRB.MINIMIZE)


# Run optimization engine
m.optimize()

# status = m.status
# if status == GRB.Status.OPTIMAL:
#     for v in m.getVars():
#         print('%s %g' % (v.varName, v.x))
#     print('Obj: %g' % m.objVal)
       
# elif status == GRB.Status.INFEASIBLE:
#     print('Optimization was stopped with status %d' % status)
#     # do IIS
#     m.computeIIS()
#     m.write("m.ilp")
#     gp.read("m.ilp")
#     for c in m.getConstrs():
#         if c.IISConstr:
#             print('%s' % c.constrName)

#Write to csv
# varInfo = [(v.varName, v.X) for v in m.getVars()]
# with open('New Variable.csv', 'w', newline='') as myfile:
#     wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
#     wr.writerows(varInfo)

# # export variables to excel
# var_names = []
# var_values = []

# for var in m.getVars():
#     if var.X > 0: 
#         var_names.append(str(var.varName))
#         var_values.append(var.X)

# # Write to csv
# with open('New Pipe Size.csv', 'w', newline='') as myfile:
#     wr = csv.writer(myfile, quoting=csv.QUOTE_ALL)
#     wr.writerows(zip(var_names, var_values))