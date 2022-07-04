import xlrd
import os
import openpyxl
import pandas as pd
from openpyxl import load_workbook
from gurobipy import *
import gurobipy as gp
from gurobipy import GRB
from itertools import product
import math

os.chdir("C:/Users/12757/Desktop/Columbia/M.S. Thesis/Optimization")

wb = load_workbook(filename='24hr quality.xlsx')
ws = wb.active

quality = {}
for column in list(ws.columns)[1:]:
    quality [column[0].value] = [(c.value) for c in column[1:]][0]

# print(quality)
# print(quality["59146264"])
# df = pd.DataFrame(quality,index=[0]).T  # transpose to look just like the sheet above
# df.to_excel('file.xls')

# wb = load_workbook(filename='24hr flowrate.xlsx')
# ws = wb.active

# flowrate = {}
# for column in list(ws.columns)[1:]:
#     flowrate [column[1].value,column[2].value] = [(c.value) for c in column[3:]][0]

# # df = pd.DataFrame(flowrate,index=[0]).T  # transpose to look just like the sheet above
# # df.to_excel('file.xls')
# # print(flowrate)
# # print(flowrate[(59104431, 'reservoir')])

# # combine positive and newnegative dictionaries
# def Merge(positive, newnagative): 
#     return(newnagative.update(positive)) 

# # extract negative flowrate into a new dic
# negative = dict((k, v) for k, v in flowrate.items() if v < 0)

# positive = dict((k, v) for k, v in flowrate.items() if v > 0)

# x = []
# y = []
# v1 = []

# for k,v in flowrate.items():
#     if v < 0:
#         x.append(k[1])
#         y.append(k[0])
#         # v1.append(v[0])
#         # new = dict((merge(x,y), -v)
#         v1.append(-v)

# new_negative = dict(zip(zip(x, y),v1))

# Merge(positive, new_negative)
# new_flowrate = new_negative 

wb = load_workbook(filename='flowrate length.xlsx')
ws = wb.active

flowrate = {}
for column in list(ws.columns)[1:]:
    flowrate [column[1].value,column[2].value,column[3].value] = [(c.value) for c in column[4:]][0]

x_pos = []
x_neg = []
y_pos = []
y_neg = []
v_pos = []
v_neg = []
length_pos = []
length_neg = []
for k,v in flowrate.items():
    if v>0:
        x_pos.append(k[0])
        y_pos.append(k[1])
        v_pos.append(v)
        length_pos.append(k[2])
    if v<0:
        x_neg.append(k[1])
        y_neg.append(k[0])
        # v1.append(v[0])
        # new = dict((merge(x,y), -v)
        v_neg.append(-v)
        length_neg.append(k[2])

x_new = x_pos + x_neg
y_new = y_pos + y_neg
v_new = v_pos + v_neg
length_new = length_pos + length_neg

new_flowrate = dict(zip(zip(x_new, y_new),v_new))
length = dict(zip(zip(x_new, y_new),length_new))
# print(new_flowrate)
# df = pd.DataFrame(new_flowrate,index=[0]).T  # transpose to look just like the sheet above
# df.to_excel('new flowrate.xls')


wb = load_workbook(filename='elevation.xlsx')
ws = wb.active

elevation = {}
for column in list(ws.columns)[1:]:
    elevation [column[0].value] = [(c.value) for c in column[1:]][0]

# wb = load_workbook(filename='length.xlsx')
# ws = wb.active

# length = {}
# for column in list(ws.columns)[1:]:
#     length [column[1].value,column[2].value] = [(c.value) for c in column[3:]][0]

# #create a list with node
# node = pd.read_csv('node.csv', header=None)
# node.values.T[0].tolist()

# node = list(node[0])

# node[269] = 1

# for i in range(0, len(node)):
#     node[i] = int(node[i])

# node[269] ='reservoir'



# link = list(flowrate.keys())
# facility = []
# node = []
# for i in range(len(link)):
#     facility.append(link[i][1])
#     node.append(link[i][0])

# # create a model
# m = gp.Model('Facility Placement')

# # create a list with facility
# facility = [59101483, 59104418, 59097687, 59082686, 59140961]

link,flowrate = gp.multidict(new_flowrate)
node,elevation = gp.multidict(elevation)
# print(node)
# print(link)

# # create variables
# x = m.addVars(facility,node,vtype=GRB.BINARY, name="x")
# y = m.addVars(facility,vtype=GRB.BINARY, name="y")
# cap = m.addVars(facility,vtype=GRB.CONTINUOUS, name="cap")

# # set parameter value
# # gpd for D, CAP_MAX, CAP_MIN,    $ for FC  $/gpd for TC
# D = []
# CAP_MAX = []
# CAP_MIN = []
# FC = [100000, 110000, 120000, 130000, 140000]
# TC = [100, 90, 80, 70, 60]

# for i in range(290):
#     D.append(22800)
    
# for i in range(5):
#     CAP_MAX.append(5000000)
#     CAP_MIN.append(0)
#     # FC.append(100000)
#     # TC.append(100)
    
# D = dict(zip(node,D))
# CAP_MAX = dict(zip(facility,CAP_MAX))
# CAP_MIN = dict(zip(facility,CAP_MIN))
# FC = dict(zip(facility,FC))
# TC = dict(zip(facility,TC))
# cl =  {key: value * 1000 for key, value in quality.items()}

# # create constraints
# m.addConstrs((gp.quicksum(x[i,j] for i in facility) == 1 for j in node), name='Demand')
# m.addConstrs((gp.quicksum(D[j] * x[i,j] for j in node) <= cap[i] for i in facility), 
#                 name='Facility Capacity') 
# m.addConstr((gp.quicksum(D[j] for j in node) == gp.quicksum(cap[i] for i in facility)), 
#                 name='Equal Total Capacity and Demand')
# m.addConstrs((CAP_MAX[i] * y[i] >= cap[i] for i in facility), name='MAX Facility Capacity')
# m.addConstrs((cap[i] >= CAP_MIN[i] * y[i] for i in facility), name='MIN Facility Capacity')

# # create objective
# # link = list(flowrate.keys())
# # facility = []
# # node = []
# # for i in range(len(link)):
# #     facility.append(link[i][1])
# #     node.append(link[i][0])
# # m.setObjective((gp.quicksum(x[i,j] * cl[i,j] for i,j in link)),GRB.MAXIMIZE) 
# # m.setObjective((gp.quicksum(x[i,j] * 1 for i,j in link)),GRB.MAXIMIZE)   
# m.setObjective((gp.quicksum(y[i] * FC[i] for i in facility) 
#                 + gp.quicksum(cap[i] * TC[i]  for i in facility)),GRB.MINIMIZE)

# # Run optimization engine
# m.optimize()
       
# #Analysis
# facility_placement = pd.DataFrame(columns=["Node", "Facility"])

# count = 0

# for j in node:
#     for i in facility:
#         if(x[i,j].x > 0.5):
#             count += 1
#             facility_placement = facility_placement.append({"Node": j, "Facility": i }, ignore_index=True )
# facility_placement.index=['']*count
# print(facility_placement)

# status = m.status
# if status == GRB.Status.OPTIMAL:
#     for v in m.getVars():
#         print('%s %g' % (v.varName, v.x))
#     print('Obj: %g' % m.objVal)
       
# elif status == GRB.Status.INFEASIBLE:
#     print('Optimization was stopped with status %d' % status)
#     # do IIS
#     m.computeIIS()
#     m.write("m.ilp")
#     gp.read("m.ilp")
#     for c in m.getConstrs():
#         if c.IISConstr:
#             print('%s' % c.constrName)

m = gp.Model('Pipe Sizing')

dup_p_in = []
dup_p_out = []
for i in range(len(link)):
    dup_p_out.append(link[i][1])
    dup_p_in.append(link[i][0])
p_in = []
p_out = []
for i in dup_p_in:
    if i not in p_in:
        p_in.append(i)
        
for i in dup_p_out:
    if i not in p_out:
        p_out.append(i)

# set parameter value
D = [0.05, 0.06, 0.08, 0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45]
Diameter = {0.05:0.05, 0.06:0.05, 0.08:0.08, 0.10:0.10, 0.15:0.15, 0.20:0.20, 
            0.25:0.25, 0.30:0.30, 0.35:0.35, 0.40:0.40, 0.45:0.45}
L = length
S_min = 0.001
S_max = 0.1
P_max = 100
Q = new_flowrate
V_min = 0.6
V_max = 3
EL = elevation 
W = 0.5
CE = 100000
CB = 10000
CPS = 10000
CP = {0.05:50, 0.06:60, 0.08:80, 0.10:100, 0.15:150, 0.20:200, 
      0.25:250, 0.30:300, 0.35:350, 0.40:400, 0.45:450}
PS_OM = 10000
COL_OM = 1000
N = []

for i in range(290):
    N.append(2)
    
N = dict(zip(node,N))

# create variables
d = m.addVars(link, D, vtype=GRB.BINARY, name="d")
e_in = m.addVars(node, vtype=GRB.CONTINUOUS, name="e_in")
e_out = m.addVars(node, vtype=GRB.CONTINUOUS, name="e_out")
p = m.addVars(link, vtype=GRB.BINARY, name="p")
pc = m.addVars(link,vtype=GRB.CONTINUOUS, name="pc")
h = m.addVars(node,vtype=GRB.CONTINUOUS, name="h")

# create constraints
m.addConstrs((gp.quicksum(d[i,j,k] for k in D) == 1 for i,j in link), name='Pipe Size')
m.addConstrs((h[i] == e_out[i]-e_in[i] for i in node), name='Elevation Change')
m.addConstrs((h[i] <= p[i,j] for i,j in link), name='Pump')
m.addConstrs((S_max >= (e_in[i]-e_in[j]+h[i])/L[i,j] for i,j in link), name='Slope LHS')
m.addConstrs(((e_in[i]-e_in[j]+h[i])/L[i,j] >= S_min for i,j in link), name='Slope RHS')
m.addConstrs((gp.quicksum(math.pi/8 * Diameter[k] ** 2 * d[i,j,k] for k in D) <= 
              V_max * Q[i,j] for i,j in link), name='Velocity')
m.addConstrs((Q[i,j] * p[i,j] <= pc[i,j] for i,j in link), name='Pump Capacity')
m.addConstr((gp.quicksum(p[i,j] for i,j in link) <= P_max ), name='Max Pump Number')

# create objective
obj1 = gp.quicksum((CE * 0.5 * ((EL[i] - e_in[i]) + (EL[j] - e_out[j])) * L[i,j] *
                   (gp.quicksum(Diameter[k] * d[i,j,k] for k in D) + 2 * W)+
                   CB * L[i,j] * (gp.quicksum(Diameter[k] * d[i,j,k] for k in D) + 2 * W)) 
                   for i,j in link)
obj2 = gp.quicksum(L[i,j] * gp.quicksum(CP[k] * Diameter[k] * d[i,j,k] for k in D) 
                   for i,j in link)
obj3 = gp.quicksum(CPS * p[i,j] for i,j in link)
obj4 = gp.quicksum(COL_OM * N[i] for i in node) + gp.quicksum(PS_OM * p[i,j] for i,j in link)
obj = obj1 + obj2 + obj3 + obj4

m.setObjective((obj),GRB.MINIMIZE)


# Run optimization engine
m.optimize()

# status = m.status
# if status == GRB.Status.OPTIMAL:
#     for v in m.getVars():
#         print('%s %g' % (v.varName, v.x))
#     print('Obj: %g' % m.objVal)
       
# elif status == GRB.Status.INFEASIBLE:
#     print('Optimization was stopped with status %d' % status)
#     # do IIS
#     m.computeIIS()
#     m.write("m.ilp")
#     gp.read("m.ilp")
#     for c in m.getConstrs():
#         if c.IISConstr:
#             print('%s' % c.constrName)


